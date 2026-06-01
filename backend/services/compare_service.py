"""
Servicio de reconciliación entre dos archivos Excel de documentos.
Compara los datos emparejando por número de cédula y reporta discrepancias.

Reconciliation service between two document Excel files.
Compares data by matching document numbers and reports discrepancies.
"""

import io
import re
import logging
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass, field
from collections import defaultdict

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# === Estilos ===
HEADER_FILL = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
NORMAL_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2E4057")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)
GREEN_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
RED_FILL = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

# Campos a comparar (clave interna -> nombre legible)
COMPARE_FIELDS = [
    ("tipo_documento", "Tipo de Documento"),
    ("nombres", "Nombres"),
    ("apellidos", "Apellidos"),
    ("fecha_nacimiento", "Fecha de Nacimiento"),
    ("sexo", "Sexo"),
    ("nacionalidad", "Nacionalidad"),
]

# Mapeo de columnas del Excel a claves internas
COLUMN_MAP = {
    "no.": "no",
    "tipo de documento": "tipo_documento",
    "numero de documento": "numero_documento",
    "nombres": "nombres",
    "apellidos": "apellidos",
    "fecha de nacimiento": "fecha_nacimiento",
    "sexo": "sexo",
    "nacionalidad": "nacionalidad",
    "estado": "estado",
    "pagina origen": "pagina_origen",
    "confianza": "confianza",
}


@dataclass
class FieldComparison:
    """Resultado de comparar un campo entre dos registros."""
    field_name: str
    value_a: str
    value_b: str
    matches: bool


@dataclass
class RecordComparison:
    """Resultado de comparar dos registros emparejados."""
    document_number: str
    fields: List[FieldComparison] = field(default_factory=list)
    all_match: bool = True
    mismatches: int = 0


@dataclass
class ReconciliationResult:
    """Resultado completo de la reconciliación."""
    matched: List[RecordComparison] = field(default_factory=list)
    only_in_a: List[Dict] = field(default_factory=list)
    only_in_b: List[Dict] = field(default_factory=list)
    stats: Dict = field(default_factory=dict)


def _clean_value(val) -> str:
    """
    Limpia un valor de celda de Excel: elimina NaN, floats .0, etc.
    Cleans an Excel cell value: removes NaN, float .0, etc.
    """
    if pd.isna(val):
        return ""
    val_str = str(val).strip()
    # Eliminar sufijo .0 de números almacenados como float en Excel
    if val_str.endswith(".0") and val_str != ".0":
        try:
            # Verificar que es realmente un float antes de quitar .0
            float(val_str)
            val_str = val_str[:-2]
        except ValueError:
            pass
    # Eliminar valores "nan" que pandas genera a veces
    if val_str.lower() == "nan":
        return ""
    return val_str


def _normalize_col_name(col: str) -> str:
    """
    Normaliza un nombre de columna: minúsculas, sin acentos, sin espacios extra.
    Normalizes a column name: lowercase, no accents, no extra spaces.
    """
    col = col.strip().lower()
    # Quitar acentos para matching tolerante
    replacements = {'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u', 'ü': 'u', 'ñ': 'n'}
    for accented, plain in replacements.items():
        col = col.replace(accented, plain)
    # Colapsar espacios múltiples
    return re.sub(r'\s+', ' ', col)


def parse_excel_for_comparison(file_bytes: bytes) -> List[Dict]:
    """
    Lee un Excel y extrae los datos normalizados para comparación.

    Reads an Excel and extracts normalized data for comparison.
    """
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="Documentos")
    except Exception:
        # Si no tiene hoja "Documentos", intentar la primera hoja
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0)

    # Normalizar nombres de columnas (sin acentos, sin espacios extra)
    df.columns = [_normalize_col_name(col) for col in df.columns]

    logger.info(f"Columnas detectadas en Excel: {list(df.columns)}")
    logger.info(f"Total filas: {len(df)}")

    # También normalizar las claves de COLUMN_MAP para el matching
    normalized_map = {_normalize_col_name(k): v for k, v in COLUMN_MAP.items()}

    # Mapear columnas
    records = []
    for _, row in df.iterrows():
        record = {}
        for col_name, key in normalized_map.items():
            if col_name in df.columns:
                record[key] = _clean_value(row.get(col_name, ""))
            else:
                record[key] = ""

        # Normalizar número de documento para emparejamiento
        record["_doc_normalized"] = _normalize_doc_number(record.get("numero_documento", ""))

        records.append(record)

    # Log de diagnóstico: mostrar cuántos registros tienen número de documento
    with_doc = sum(1 for r in records if r.get("_doc_normalized"))
    logger.info(f"Registros parseados: {len(records)}, con numero_documento: {with_doc}")

    # Log de los primeros 3 documentos normalizados para diagnóstico
    for i, r in enumerate(records[:3]):
        logger.debug(
            f"  Registro {i+1}: doc='{r.get('numero_documento', '')}' "
            f"-> normalized='{r.get('_doc_normalized', '')}'"
        )

    return records


def _normalize_doc_number(doc: str) -> str:
    """
    Normaliza un número de documento quitando puntos, espacios y guiones.
    Normalizes a document number by removing dots, spaces, and hyphens.
    """
    if not doc:
        return ""
    return re.sub(r'[.\s\-]', '', doc).strip()


def _normalize_for_comparison(value: str, field_name: str) -> str:
    """
    Normaliza un valor para comparación tolerante.
    Normaliza según el tipo de campo.

    Normalizes a value for tolerant comparison.
    Normalizes based on field type.
    """
    if not value:
        return ""

    value = value.strip().lower()

    if field_name == "fecha_nacimiento":
        return _normalize_date(value)
    elif field_name == "sexo":
        return _normalize_sex(value)
    else:
        # Quitar acentos para comparación tolerante
        return _remove_accents(value)


def _normalize_date(date_str: str) -> str:
    """
    Normaliza una fecha a formato YYYYMMDD para comparación.
    Normalizes a date to YYYYMMDD format for comparison.
    """
    if not date_str:
        return ""

    date_str = date_str.strip()

    # DD/MM/YYYY
    match = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', date_str)
    if match:
        return f"{match.group(3)}{match.group(2).zfill(2)}{match.group(1).zfill(2)}"

    # YYYY-MM-DD
    match = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', date_str)
    if match:
        return f"{match.group(1)}{match.group(2).zfill(2)}{match.group(3).zfill(2)}"

    # DD-MM-YYYY
    match = re.match(r'^(\d{1,2})-(\d{1,2})-(\d{4})$', date_str)
    if match:
        return f"{match.group(3)}{match.group(2).zfill(2)}{match.group(1).zfill(2)}"

    return date_str.lower()


def _normalize_sex(sex: str) -> str:
    """
    Normaliza el campo sexo.
    Normalizes the sex field.
    """
    sex = sex.strip().lower()
    if sex in ("m", "masculino", "male"):
        return "M"
    elif sex in ("f", "femenino", "female"):
        return "F"
    return sex


def _remove_accents(text: str) -> str:
    """
    Quita acentos para comparación tolerante.
    Removes accents for tolerant comparison.
    """
    replacements = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
        'ü': 'u', 'ñ': 'n',
    }
    for accented, plain in replacements.items():
        text = text.replace(accented, plain)
    return text


def reconcile(file_a_data: List[Dict], file_b_data: List[Dict]) -> ReconciliationResult:
    """
    Reconcilia dos conjuntos de datos emparejando por número de documento.
    Compara campo a campo con tolerancia (mayúsculas, acentos, formato fecha).

    Reconciles two datasets by matching document numbers.
    Compares field by field with tolerance (case, accents, date format).
    """
    result = ReconciliationResult()

    # Indexar por número de documento normalizado (listas para soportar duplicados)
    index_a = defaultdict(list)
    for record in file_a_data:
        doc = record.get("_doc_normalized", "")
        if doc:
            index_a[doc].append(record)
        else:
            # Registros sin número de documento van directo a only_in_a
            result.only_in_a.append(record)

    index_b = defaultdict(list)
    for record in file_b_data:
        doc = record.get("_doc_normalized", "")
        if doc:
            index_b[doc].append(record)
        else:
            # Registros sin número de documento van directo a only_in_b
            result.only_in_b.append(record)

    logger.info(
        f"Índices: A tiene {len(index_a)} cédulas únicas, "
        f"B tiene {len(index_b)} cédulas únicas"
    )
    # Log primeros 5 docs de cada índice para diagnóstico
    logger.debug(f"  Docs A (muestra): {list(index_a.keys())[:5]}")
    logger.debug(f"  Docs B (muestra): {list(index_b.keys())[:5]}")

    # Emparejar
    total_fields_compared = 0
    total_matches = 0
    field_mismatch_counts = {field_key: 0 for field_key, _ in COMPARE_FIELDS}

    for doc_num, records_a in index_a.items():
        records_b = index_b.get(doc_num, [])
        if records_b:
            # Emparejar uno a uno, consumiendo de ambas listas
            pairs = min(len(records_a), len(records_b))
            for i in range(pairs):
                record_a = records_a[i]
                record_b = records_b[i]

                comparison = RecordComparison(
                    document_number=record_a.get("numero_documento", doc_num)
                )

                for field_key, field_label in COMPARE_FIELDS:
                    val_a = record_a.get(field_key, "")
                    val_b = record_b.get(field_key, "")

                    norm_a = _normalize_for_comparison(val_a, field_key)
                    norm_b = _normalize_for_comparison(val_b, field_key)

                    total_fields_compared += 1

                    matches = (norm_a == norm_b) or (not norm_a and not norm_b)

                    if matches:
                        total_matches += 1
                    else:
                        comparison.all_match = False
                        comparison.mismatches += 1
                        field_mismatch_counts[field_key] += 1

                    comparison.fields.append(FieldComparison(
                        field_name=field_label,
                        value_a=val_a,
                        value_b=val_b,
                        matches=matches,
                    ))

                result.matched.append(comparison)

            # Sobrantes en A (más registros en A que en B para esta cédula)
            for i in range(pairs, len(records_a)):
                result.only_in_a.append(records_a[i])
            # Sobrantes en B (más registros en B que en A para esta cédula)
            for i in range(pairs, len(records_b)):
                result.only_in_b.append(records_b[i])
        else:
            # Todos los registros con esta cédula solo están en A
            result.only_in_a.extend(records_a)

    # Registros solo en B (cédulas que no existen en A)
    for doc_num, records_b in index_b.items():
        if doc_num not in index_a:
            result.only_in_b.extend(records_b)

    # Estadísticas
    total_records = len(file_a_data) + len(file_b_data)
    total_matched = len(result.matched)
    total_only_a = len(result.only_in_a)
    total_only_b = len(result.only_in_b)
    total_mismatches = sum(1 for m in result.matched if not m.all_match)

    result.stats = {
        "total_records_a": len(file_a_data),
        "total_records_b": len(file_b_data),
        "matched_pairs": total_matched,
        "only_in_a": total_only_a,
        "only_in_b": total_only_b,
        "total_fields_compared": total_fields_compared,
        "matching_fields": total_matches,
        "mismatching_fields": total_fields_compared - total_matches,
        "records_with_mismatches": total_mismatches,
        "records_all_match": total_matched - total_mismatches,
        "accuracy_pct": round((total_matches / total_fields_compared * 100), 1) if total_fields_compared > 0 else 0,
        "field_mismatch_counts": field_mismatch_counts,
        "all_clear": total_mismatches == 0 and total_only_a == 0 and total_only_b == 0,
    }

    return result


def create_reconciliation_excel(result: ReconciliationResult) -> io.BytesIO:
    """
    Genera un Excel con el reporte de reconciliación.

    Generates an Excel with the reconciliation report.
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _write_reconciliation_sheet(writer, result)
        _write_only_in_sheet(writer, result.only_in_a, "Solo en Archivo A", "🔴")
        _write_only_in_sheet(writer, result.only_in_b, "Solo en Archivo B", "🔵")
        _write_summary_sheet(writer, result)

    output.seek(0)
    return output


def _write_reconciliation_sheet(writer, result: ReconciliationResult):
    """
    Escribe la hoja de conciliación con comparación campo a campo.

    Writes the reconciliation sheet with field-by-field comparison.
    """
    rows = []
    row_num = 0

    for comparison in result.matched:
        row_num += 1
        for fc in comparison.fields:
            rows.append([
                row_num,
                comparison.document_number,
                fc.field_name,
                fc.value_a or "(vacío)",
                fc.value_b or "(vacío)",
                "✅ Coincide" if fc.matches else "❌ Difiere",
            ])

    if not rows:
        rows = [["", "", "", "", "", "Sin datos para comparar"]]

    columns = ["No.", "Cédula", "Campo", "Archivo A", "Archivo B", "Estado"]
    df = pd.DataFrame(rows, columns=columns)
    df.to_excel(writer, index=False, sheet_name="Conciliación")

    ws = writer.sheets["Conciliación"]

    # Encabezados
    for col in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN

    # Formato de datos
    for row_idx in range(2, len(rows) + 2):
        status_cell = ws.cell(row=row_idx, column=6)
        status_val = str(status_cell.value or "")

        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN

            if "Difiere" in status_val:
                # Fila con discrepancia: rojo suave
                if col_idx in (4, 5):  # Columnas de valores
                    cell.fill = RED_FILL
            elif "Coincide" in status_val:
                if col_idx == 6:
                    cell.fill = GREEN_FILL

        status_cell.alignment = CENTER_ALIGN

    # Anchos
    widths = [6, 22, 22, 25, 25, 16]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"
    ws.freeze_panes = "A2"


def _write_only_in_sheet(writer, records: List[Dict], sheet_name: str, emoji: str):
    """
    Escribe la hoja de registros que solo están en un archivo.

    Writes the sheet for records that are only in one file.
    """
    if not records:
        return

    display_cols = ["numero_documento", "nombres", "apellidos", "fecha_nacimiento", "sexo", "nacionalidad"]
    col_labels = ["Número de Documento", "Nombres", "Apellidos", "Fecha de Nacimiento", "Sexo", "Nacionalidad"]

    rows = []
    for record in records:
        rows.append([record.get(c, "") for c in display_cols])

    df = pd.DataFrame(rows, columns=col_labels)
    df.to_excel(writer, index=False, sheet_name=sheet_name)

    ws = writer.sheets[sheet_name]

    for col in range(1, len(col_labels) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN

    for row_idx in range(2, len(rows) + 2):
        for col_idx in range(1, len(col_labels) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

    widths = [22, 25, 25, 18, 8, 16]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    ws.freeze_panes = "A2"


def _write_summary_sheet(writer, result: ReconciliationResult):
    """
    Escribe la hoja de resumen con estadísticas.

    Writes the summary sheet with statistics.
    """
    ws = writer.book.create_sheet("Resumen")
    stats = result.stats

    # Título
    title = "✅ Sin incongruencias — Todos los datos coinciden" if stats["all_clear"] \
        else f"⚠️ {stats['mismatching_fields']} incongruencias encontradas"
    ws.cell(row=1, column=1, value="Resumen de Conciliación")
    ws.cell(row=1, column=1).font = TITLE_FONT

    ws.cell(row=2, column=1, value=title)
    ws.cell(row=2, column=1).font = BOLD_FONT
    if stats["all_clear"]:
        ws.cell(row=2, column=1).fill = GREEN_FILL
    else:
        ws.cell(row=2, column=1).fill = YELLOW_FILL

    # Métricas generales
    ws.cell(row=4, column=1, value="Métricas Generales")
    ws.cell(row=4, column=1).font = BOLD_FONT

    metrics = [
        ("Total registros Archivo A", stats["total_records_a"]),
        ("Total registros Archivo B", stats["total_records_b"]),
        ("Registros emparejados", stats["matched_pairs"]),
        ("Registros solo en A", stats["only_in_a"]),
        ("Registros solo en B", stats["only_in_b"]),
        ("Campos comparados", stats["total_fields_compared"]),
        ("Campos coincidentes", stats["matching_fields"]),
        ("Campos con diferencias", stats["mismatching_fields"]),
        ("Registros con discrepancias", stats["records_with_mismatches"]),
        ("Registros sin discrepancias", stats["records_all_match"]),
        ("Precisión global", f"{stats['accuracy_pct']}%"),
    ]

    for i, (label, value) in enumerate(metrics):
        row = 5 + i
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        cell_val = ws.cell(row=row, column=2, value=value)
        cell_val.font = BOLD_FONT
        cell_val.border = THIN_BORDER
        cell_val.alignment = CENTER_ALIGN

        # Colorear precisión
        if label == "Precisión global" and isinstance(value, str):
            pct = float(value.replace("%", ""))
            if pct >= 95:
                cell_val.fill = GREEN_FILL
            elif pct >= 80:
                cell_val.fill = YELLOW_FILL
            else:
                cell_val.fill = RED_FILL

    # Desglose por campo
    ws.cell(row=18, column=1, value="Discrepancias por Campo")
    ws.cell(row=18, column=1).font = BOLD_FONT

    headers = ["Campo", "Discrepancias"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=19, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN

    row = 20
    for field_key, field_label in COMPARE_FIELDS:
        mismatches = stats["field_mismatch_counts"].get(field_key, 0)
        ws.cell(row=row, column=1, value=field_label).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        cell_val = ws.cell(row=row, column=2, value=mismatches)
        cell_val.font = NORMAL_FONT
        cell_val.border = THIN_BORDER
        cell_val.alignment = CENTER_ALIGN
        cell_val.fill = RED_FILL if mismatches > 0 else GREEN_FILL
        row += 1

    # Anchos
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 14
