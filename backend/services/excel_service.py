"""
Servicio para crear archivos Excel con los datos extraidos de documentos colombianos.
Excel profesional con formato condicional, filtros y resumen detallado.

Service for creating Excel files with data extracted from Colombian documents.
Professional Excel with conditional formatting, filters, and detailed summary.
"""

import pandas as pd
import io
import logging
from typing import List, Dict, Optional
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Columnas principales del Excel
# Main Excel columns
EXCEL_COLUMNS = [
    "No.",
    "Tipo de Documento",
    "Numero de Documento",
    "Nombres",
    "Apellidos",
    "Fecha de Nacimiento",
    "Sexo",
    "Nacionalidad",
    "Estado",
    "Pagina Origen",
    "Confianza"
]

# Columnas del reporte de problemas
# Problem report columns
PROBLEM_COLUMNS = [
    "Pagina",
    "Estado",
    "Tipo de Cara",
    "Tipo de Documento",
    "Problema",
    "Accion Sugerida"
]

# Estilos
# Styles
HEADER_FILL = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
NORMAL_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2E4057")

RED_FILL = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
GREEN_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _is_document_complete(item: Dict[str, str]) -> bool:
    """Determina si un documento tiene todos los campos criticos.
    Determines whether a document has all critical fields."""
    critical = ["numero_documento", "nombres", "apellidos"]
    return all(item.get(f, "").strip() for f in critical)


def _get_confidence_label(confidence: float) -> str:
    """Convierte la confianza a etiqueta legible.
    Converts confidence to a readable label."""
    if confidence >= 0.9:
        return "Alta"
    elif confidence >= 0.7:
        return "Media"
    elif confidence >= 0.5:
        return "Baja"
    else:
        return "Muy Baja"


def create_excel_with_merged_documents(
    documents: List[Dict[str, str]],
    problems: Optional[List[Dict[str, str]]] = None,
    page_results: Optional[List] = None
) -> io.BytesIO:
    """
    Crea un archivo Excel profesional con formato condicional.

    Creates a professional Excel file with conditional formatting.
    """
    try:
        excel_buffer = io.BytesIO()

        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            _write_documents_sheet(writer, documents)
            _write_problems_sheet(writer, problems)
            _write_summary_sheet(writer, documents, problems)

        excel_buffer.seek(0)
        return excel_buffer

    except Exception as e:
        logger.error(f"Error al crear Excel: {e}")
        return _create_fallback_excel(documents)


def _write_documents_sheet(writer, documents: List[Dict[str, str]]):
    """Escribe la hoja principal de documentos con formato.
    Writes the main documents sheet with formatting."""
    rows = []
    for idx, item in enumerate(documents, 1):
        is_complete = _is_document_complete(item)
        rows.append([
            idx,
            item.get("tipo_documento", ""),
            item.get("numero_documento", ""),
            item.get("nombres", ""),
            item.get("apellidos", ""),
            item.get("fecha_nacimiento", ""),
            item.get("sexo", ""),
            item.get("nacionalidad", ""),
            "Completo" if is_complete else "Incompleto",
            item.get("_page_origin", ""),
            item.get("_confidence_label", ""),
        ])

    df = pd.DataFrame(rows, columns=EXCEL_COLUMNS)
    df.to_excel(writer, index=False, sheet_name="Documentos")

    ws = writer.sheets["Documentos"]

    # Formato de encabezados
    # Header formatting
    for col in range(1, len(EXCEL_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN

    # Formato de datos
    # Data formatting
    for row in range(2, len(documents) + 2):
        item = documents[row - 2]
        is_complete = _is_document_complete(item)

        for col in range(1, len(EXCEL_COLUMNS) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN

            # Filas alternas
            # Alternating rows
            if row % 2 == 0:
                cell.fill = LIGHT_GRAY_FILL

            value = cell.value
            if value is None:
                value = ""

        # Celda de estado con color
        # Status cell with color
        status_col = EXCEL_COLUMNS.index("Estado") + 1
        status_cell = ws.cell(row=row, column=status_col)
        status_cell.alignment = CENTER_ALIGN
        if is_complete:
            status_cell.fill = GREEN_FILL
            status_cell.font = Font(name="Calibri", bold=True, color="065F46", size=11)
        else:
            status_cell.fill = RED_FILL
            status_cell.font = Font(name="Calibri", bold=True, color="991B1B", size=11)

        # Confianza con color
        # Confidence with color
        conf_col = EXCEL_COLUMNS.index("Confianza") + 1
        conf_cell = ws.cell(row=row, column=conf_col)
        conf_cell.alignment = CENTER_ALIGN
        conf_val = conf_cell.value
        if conf_val == "Alta":
            conf_cell.fill = GREEN_FILL
        elif conf_val == "Baja" or conf_val == "Muy Baja":
            conf_cell.fill = YELLOW_FILL

        # Celdas vacias en rojo claro (solo columnas de datos, no No./Estado/etc)
        # Empty cells in light red (data columns only, not No./Status/etc)
        data_cols = [2, 3, 4, 5, 6, 7, 8]  # Tipo Doc, Numero, Nombres, Apellidos, Fecha, Sexo, Nac
        for col in data_cols:
            cell = ws.cell(row=row, column=col)
            if cell.value is None or str(cell.value).strip() == "":
                cell.fill = RED_FILL

    # Ancho de columnas
    # Column widths
    col_widths = {
        "No.": 6, "Tipo de Documento": 28, "Numero de Documento": 22,
        "Nombres": 25, "Apellidos": 25, "Fecha de Nacimiento": 18,
        "Sexo": 8, "Nacionalidad": 16, "Estado": 14,
        "Pagina Origen": 14, "Confianza": 12
    }
    for col_name, width in col_widths.items():
        col_idx = EXCEL_COLUMNS.index(col_name) + 1
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Filtros automaticos
    # Auto filters
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLUMNS))}{len(documents) + 1}"

    # Congelar primera fila
    # Freeze first row
    ws.freeze_panes = "A2"


def _write_problems_sheet(writer, problems: Optional[List[Dict[str, str]]]):
    """Escribe la hoja de problemas.
    Writes the problems sheet."""
    if not problems:
        return

    rows = []
    for p in problems:
        rows.append([
            p.get("Pagina", ""),
            p.get("Estado", ""),
            p.get("Tipo de Cara", ""),
            p.get("Tipo de Documento", ""),
            p.get("Problema", ""),
            p.get("Accion Sugerida", ""),
        ])

    df = pd.DataFrame(rows, columns=PROBLEM_COLUMNS)
    df.to_excel(writer, index=False, sheet_name="Revision Requerida")

    ws = writer.sheets["Revision Requerida"]

    # Encabezados
    # Headers
    for col in range(1, len(PROBLEM_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN

    # Datos con formato
    # Formatted data
    for row in range(2, len(problems) + 2):
        for col in range(1, len(PROBLEM_COLUMNS) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if row % 2 == 0:
                cell.fill = LIGHT_GRAY_FILL

        # Estado con color
        # Status with color
        status_cell = ws.cell(row=row, column=2)
        status_val = str(status_cell.value or "")
        if "ERROR" in status_val:
            status_cell.fill = RED_FILL
        elif "INCOMPLETO" in status_val:
            status_cell.fill = YELLOW_FILL

    # Anchos
    # Widths
    prob_widths = [10, 18, 16, 28, 40, 40]
    for i, w in enumerate(prob_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    ws.auto_filter.ref = f"A1:{get_column_letter(len(PROBLEM_COLUMNS))}{len(problems) + 1}"
    ws.freeze_panes = "A2"

    logger.info(f"Reporte de problemas: {len(problems)} paginas requieren revision")


def _write_summary_sheet(writer, documents: List[Dict[str, str]], problems: Optional[List[Dict[str, str]]]):
    """Escribe la hoja de resumen con estadisticas.
    Writes the summary sheet with statistics."""
    ws = writer.book.create_sheet("Resumen")

    total = len(documents)
    completos = sum(1 for d in documents if _is_document_complete(d))
    incompletos = total - completos
    errors = sum(1 for p in (problems or []) if "ERROR" in p.get("Estado", ""))
    incomplete_pages = sum(1 for p in (problems or []) if "INCOMPLETO" in p.get("Estado", ""))

    # Titulo
    # Title
    ws.cell(row=1, column=1, value="Resumen de Extraccion")
    ws.cell(row=1, column=1).font = TITLE_FONT

    # Metricas generales
    # General metrics
    ws.cell(row=3, column=1, value="Metricas Generales")
    ws.cell(row=3, column=1).font = BOLD_FONT

    metrics = [
        ("Total documentos extraidos", total),
        ("Documentos completos", completos),
        ("Documentos incompletos", incompletos),
        ("Paginas con errores", errors),
        ("Paginas con datos incompletos", incomplete_pages),
        ("Tasa de exito", f"{(completos / total * 100):.1f}%" if total > 0 else "N/A"),
    ]

    for i, (label, value) in enumerate(metrics):
        row = 4 + i
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        cell_val = ws.cell(row=row, column=2, value=value)
        cell_val.font = BOLD_FONT
        cell_val.border = THIN_BORDER
        cell_val.alignment = CENTER_ALIGN

        # Colorear tasa
        # Color the rate
        if label == "Tasa de exito" and isinstance(value, str):
            pct = float(value.replace("%", ""))
            if pct >= 90:
                cell_val.fill = GREEN_FILL
            elif pct >= 70:
                cell_val.fill = YELLOW_FILL
            else:
                cell_val.fill = RED_FILL

    # Desglose por tipo de documento
    # Breakdown by document type
    ws.cell(row=12, column=1, value="Desglose por Tipo de Documento")
    ws.cell(row=12, column=1).font = BOLD_FONT

    type_counts: Dict[str, int] = {}
    type_complete: Dict[str, int] = {}
    for doc in documents:
        tipo = doc.get("tipo_documento", "Sin tipo")
        type_counts[tipo] = type_counts.get(tipo, 0) + 1
        if _is_document_complete(doc):
            type_complete[tipo] = type_complete.get(tipo, 0) + 1

    headers_tipo = ["Tipo de Documento", "Cantidad", "Completos", "Incompletos"]
    for col, header in enumerate(headers_tipo, 1):
        cell = ws.cell(row=13, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN

    row = 14
    for tipo, count in sorted(type_counts.items(), key=lambda x: -x[1]):
        complete = type_complete.get(tipo, 0)
        ws.cell(row=row, column=1, value=tipo).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=count).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = CENTER_ALIGN
        ws.cell(row=row, column=3, value=complete).font = NORMAL_FONT
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=3).fill = GREEN_FILL
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN
        ws.cell(row=row, column=4, value=count - complete).font = NORMAL_FONT
        ws.cell(row=row, column=4).border = THIN_BORDER
        ws.cell(row=row, column=4).fill = RED_FILL if count - complete > 0 else GREEN_FILL
        ws.cell(row=row, column=4).alignment = CENTER_ALIGN
        row += 1

    # Anchos
    # Widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14


def _create_fallback_excel(documents: List[Dict[str, str]]) -> io.BytesIO:
    """Excel basico de emergencia si el formato completo falla.
    Basic emergency Excel if full formatting fails."""
    excel_buffer = io.BytesIO()
    df = pd.DataFrame(columns=EXCEL_COLUMNS)
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Documentos")
    excel_buffer.seek(0)
    return excel_buffer


def get_excel_columns() -> List[str]:
    return EXCEL_COLUMNS.copy()


def validate_data_for_excel(data: Dict[str, str]) -> bool:
    if not data.get("numero_documento"):
        return False
    if not data.get("tipo_documento"):
        return False
    return True
