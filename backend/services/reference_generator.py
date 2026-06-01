"""
Generador de Excel de referencia con filas mezcladas al azar.
Toma un Excel extraído y devuelve el mismo con orden aleatorio.

Reference Excel generator with randomly shuffled rows.
Takes an extracted Excel and returns it in random order.
"""

import random
import io
import logging
from typing import List, Dict

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Estilos reutilizados de excel_service.py
HEADER_FILL = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
NORMAL_FONT = Font(name="Calibri", size=11)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)
LIGHT_GRAY_FILL = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

# Columnas del Excel de documentos
DOCUMENT_COLUMNS = [
    "No.",
    "Tipo de Documento",
    "Numero de Documento",
    "Nombres",
    "Apellidos",
    "Fecha de Nacimiento",
    "Sexo",
    "Nacionalidad",
    "Estado",
    "Pagina Origen",
    "Confianza"
]


def generate_shuffled_excel(source_bytes: bytes) -> io.BytesIO:
    """
    Lee un Excel extraído, mezcla las filas al azar y genera un nuevo Excel.

    Args:
        source_bytes: Bytes del archivo Excel fuente

    Returns:
        BytesIO con el Excel mezclado

    Reads an extracted Excel, shuffles rows randomly, and generates a new Excel.

    Args:
        source_bytes: Bytes of the source Excel file

    Returns:
        BytesIO with the shuffled Excel
    """
    try:
        # Leemos la hoja "Documentos"
        df = pd.read_excel(io.BytesIO(source_bytes), sheet_name="Documentos")

        # Convertimos a lista de diccionarios para mezclar
        rows = df.to_dict("records")

        # Mezclamos al azar
        random.shuffle(rows)

        # Renumeramos la columna "No."
        for i, row in enumerate(rows, 1):
            row["No."] = i

        # Generamos el Excel con formato
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_shuffled = pd.DataFrame(rows, columns=DOCUMENT_COLUMNS)
            df_shuffled.to_excel(writer, index=False, sheet_name="Documentos")

        # Aplicamos formato
        _apply_formatting(output)

        output.seek(0)
        return output

    except Exception as e:
        logger.error(f"Error al generar Excel mezclado: {e}")
        raise


def _apply_formatting(buffer: io.BytesIO):
    """
    Aplica formato profesional al Excel generado.

    Applies professional formatting to the generated Excel.
    """
    from openpyxl import load_workbook

    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb["Documentos"]

    # Encabezados
    for col in range(1, len(DOCUMENT_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN

    # Datos con formato
    total_rows = ws.max_row - 1
    for row_idx in range(2, total_rows + 2):
        for col_idx in range(1, len(DOCUMENT_COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN

            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

    # Anchos de columna
    col_widths = {
        "No.": 6, "Tipo de Documento": 28, "Numero de Documento": 22,
        "Nombres": 25, "Apellidos": 25, "Fecha de Nacimiento": 18,
        "Sexo": 8, "Nacionalidad": 16, "Estado": 14,
        "Pagina Origen": 14, "Confianza": 12
    }
    for col_name, width in col_widths.items():
        if col_name in DOCUMENT_COLUMNS:
            col_idx = DOCUMENT_COLUMNS.index(col_name) + 1
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Filtros y congelar
    ws.auto_filter.ref = f"A1:{get_column_letter(len(DOCUMENT_COLUMNS))}{total_rows + 1}"
    ws.freeze_panes = "A2"

    # Guardamos de vuelta al buffer
    buffer.seek(0)
    buffer.truncate()
    wb.save(buffer)
